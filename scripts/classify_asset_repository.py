#!/usr/bin/env python3
"""Classify production-format files not used as an exact asset-ledger path.

This is the AST-04 evidence generator. It is deliberately read-only: root assets
must be registered in the workbook first; child exports, textures, materials,
runtime wrappers, and historical versions are associated with a stable parent
AssetID in the JSON report instead of receiving fabricated root IDs.
"""

from __future__ import annotations

import argparse
import json
import re
import subprocess
import sys
from collections import Counter
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


PRODUCTION_EXTENSIONS = {
    ".png", ".webp", ".svg", ".wav", ".ogg", ".mp3", ".jpg", ".jpeg",
    ".glb", ".gltf", ".fbx", ".tres", ".tscn", ".blend",
}
EXCLUDED_DIRECTORIES = {"source", "previews", "reports"}
GENERIC_TOKENS = {
    "assets", "art", "audio", "runtime", "components", "component", "visual",
    "root", "top3d", "game", "scene", "asset", "props", "characters",
    "environments", "weapons", "shared", "variants", "accessories", "head",
}
VERSION_PATTERN = re.compile(r"^v(\d{3})$")
PATH_SPLIT_PATTERN = re.compile(r"[;；\n]+")


def _text(value: Any) -> str:
    return "" if value is None else str(value).strip()


def _tokens(value: str) -> set[str]:
    raw = re.findall(r"[a-z0-9]+", value.lower().replace("top3d", " top3d "))
    return {
        token for token in raw
        if token not in GENERIC_TOKENS and not VERSION_PATTERN.fullmatch(token) and len(token) > 1
    }


def _registry_rows(workbook_path: Path, project_root: Path) -> tuple[list[dict[str, Any]], set[str]]:
    workbook = load_workbook(workbook_path, read_only=False, data_only=False)
    sheet = workbook["资产主表"]
    rows: list[dict[str, Any]] = []
    exact_paths: set[str] = set()
    for row_number in range(6, sheet.max_row + 1):
        asset_id = _text(sheet.cell(row_number, 1).value)
        if not asset_id:
            continue
        values = [_text(sheet.cell(row_number, column).value) for column in range(1, 25)]
        paths = []
        for token in PATH_SPLIT_PATTERN.split(values[14]):
            token = token.strip().removeprefix("res://")
            if token:
                paths.append(token)
                if (project_root / token).is_file():
                    exact_paths.add(token)
        rows.append({
            "row": row_number,
            "asset_id": asset_id,
            "paths": paths,
            "corpus": " ".join(values),
            "tokens": _tokens(" ".join(values)),
        })
    return rows, exact_paths


def _text_reference_blob(project_root: Path) -> str:
    tracked = subprocess.check_output(
        ["git", "ls-files", "src", "scenes", "tests", "assets", "project.godot"],
        cwd=project_root,
        text=True,
    ).splitlines()
    chunks: list[str] = []
    for relative in tracked:
        path = project_root / relative
        if path.suffix.lower() not in {".gd", ".tscn", ".tres", ".import", ".md", ".json"}:
            continue
        try:
            chunks.append(path.read_text(encoding="utf-8", errors="ignore"))
        except OSError:
            continue
    return "\n".join(chunks)


def _parent_for(relative_path: str, rows: list[dict[str, Any]]) -> tuple[str, int, str]:
    normalized = relative_path.lower()
    override_id = ""
    if "/shared/palette/" in normalized:
        override_id = "ART-PALETTE-FACILITY-LOWLIGHT-DOPAMINE"
    elif "/rooftop_shelter_3d/" in normalized:
        override_id = "ENV-ROOFTOP-SHELTER-50M-3D"
    elif "chibi_anime_head_" in normalized or "head_chibi_anime_" in normalized:
        override_id = "CHR-PLY-BUNNY01-HEAD-CHIBI-ANIME-3D"
    elif "chr_player_capsule01_bunny01_" in normalized:
        if "_body_" in normalized:
            override_id = "CHR-PLY-CAPSULE01-3D-BUNNY01-BODY"
        elif "_head_" in normalized:
            override_id = "CHR-PLY-CAPSULE01-3D-BUNNY01-HEAD"
        elif "_ear_" in normalized:
            override_id = "CHR-PLY-CAPSULE01-3D-BUNNY01-EARS"
        elif "_hand_" in normalized:
            override_id = "CHR-PLY-CAPSULE01-3D-BUNNY01-HAND"
        elif "_foot_l_" in normalized:
            override_id = "CHR-PLY-CAPSULE01-3D-BUNNY01-FOOT-L"
        elif "_foot_r_" in normalized:
            override_id = "CHR-PLY-CAPSULE01-3D-BUNNY01-FOOT-R"
        else:
            override_id = "CHR-PLY-CAPSULE01-3D-BUNNY01"
    else:
        dungeon_parent_by_stem = {
            "prp_room_door_3d": "PRP-TOWER-ROOM-DOOR-3D",
            "prp_tower_floor_tile_5m": "ENV-TOWER-FLOOR-TILE-5M",
            "prp_tower_stairwell_generic_9m": "ENV-TOWER-STAIRWELL-GENERIC-9M",
            "prp_tower_stairwell_rooftop_9m": "ENV-TOWER-STAIRWELL-ROOFTOP-9M",
            "prp_tower_wall_door_5m": "ENV-TOWER-WALL-DOOR-5M",
            "prp_tower_wall_parapet_5m": "ENV-TOWER-WALL-PARAPET-5M",
            "prp_tower_wall_solid_5m": "ENV-TOWER-WALL-SOLID-5M",
        }
        normalized_stem = re.sub(r"_v\d{3}$", "", Path(relative_path).stem.lower())
        override_id = dungeon_parent_by_stem.get(normalized_stem, "")
    if override_id:
        for row in rows:
            if row["asset_id"] == override_id:
                return override_id, row["row"], "explicit_asset_family"

    candidate_tokens = _tokens(relative_path)
    candidate_parts = set(Path(relative_path).parts)
    best: tuple[int, str, int, str] | None = None
    for row in rows:
        common_tokens = candidate_tokens & row["tokens"]
        shared_parts = max(
            (len(candidate_parts & set(Path(registered_path).parts)) for registered_path in row["paths"]),
            default=0,
        )
        score = len(common_tokens) * 4 + shared_parts * 2
        logical_key = Path(relative_path).stem.lower()
        for generic_suffix in ("_visual_top3d", "_root_top3d", "_game", "_color", "_normal"):
            logical_key = logical_key.split(generic_suffix)[0]
        logical_key = re.sub(r"_v\d{3}.*$", "", logical_key)
        if logical_key and logical_key in row["corpus"].lower():
            score += 24
        if score <= 0:
            continue
        candidate = (score, row["asset_id"], row["row"], ",".join(sorted(common_tokens)))
        if best is None or candidate > best:
            best = candidate
    if best is None:
        return "", 0, ""
    return best[1], best[2], best[3]


def classify(workbook_path: Path, project_root: Path) -> dict[str, Any]:
    rows, exact_paths = _registry_rows(workbook_path, project_root)
    references = _text_reference_blob(project_root)
    candidates: list[str] = []
    for path in (project_root / "assets").rglob("*"):
        if not path.is_file() or path.suffix.lower() not in PRODUCTION_EXTENSIONS:
            continue
        relative = path.relative_to(project_root).as_posix()
        if any(part in EXCLUDED_DIRECTORIES for part in path.relative_to(project_root).parts):
            continue
        if relative not in exact_paths:
            candidates.append(relative)

    records: list[dict[str, Any]] = []
    for relative in sorted(candidates):
        path = Path(relative)
        parent_id, parent_row, token_basis = _parent_for(relative, rows)
        referenced = relative in references or f"res://{relative}" in references or path.name in references
        if "chibi_anime_head_v001" in relative or "bunny01_root_top3d_v007" in relative:
            classification = "historical_or_intermediate_version"
            reason = "父 AssetID 已登记更高运行时版本；该文件保留作可追溯历史"
        elif path.suffix.lower() in {".png", ".jpg", ".jpeg", ".webp", ".tres"}:
            classification = "registered_parent_dependency"
            reason = "贴图/材质由父资产或其导入链消费，不建立根 AssetID"
        elif path.suffix.lower() == ".tscn" and referenced:
            classification = "runtime_wrapper"
            reason = "Godot 运行时包装场景；父 AssetID 保持在模型/设施根资产"
        elif referenced:
            classification = "current_export_dependency"
            reason = "当前场景或代码直接引用的导出依赖；归属父 AssetID"
        else:
            classification = "historical_or_intermediate_version"
            reason = "无当前文本引用的旧版本/中间导出；保留作可追溯历史，不作为根资产"
        records.append({
            "path": relative,
            "classification": classification,
            "parent_asset_id": parent_id,
            "parent_registry_row": parent_row,
            "referenced_by_current_text_assets": referenced,
            "classification_basis": reason,
            "parent_match_tokens": token_basis,
        })

    unresolved = [record for record in records if not record["parent_asset_id"]]
    return {
        "workbook": str(workbook_path),
        "scan_contract": {
            "extensions": sorted(PRODUCTION_EXTENSIONS),
            "excluded_directories": sorted(EXCLUDED_DIRECTORIES),
            "exact_ledger_paths_are_not_candidates": True,
        },
        "baseline_note": (
            "初检的254是启发式计数且未保存逐文件清单；AST-04以当前规范化工作树重建精确清单。"
            "8个BGM及共享色盘、角色头部配件、恢复站、房门、屋顶避难所共13个根资产已登记，"
            "当前剩余候选均在本报告逐文件归属。"
        ),
        "new_root_assets_registered_ast04": 13,
        "classified_candidate_count": len(records),
        "classification_counts": dict(Counter(record["classification"] for record in records)),
        "unresolved_parent_count": len(unresolved),
        "unresolved": unresolved,
        "records": records,
    }


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--project-root", type=Path, default=Path(__file__).resolve().parents[1])
    parser.add_argument("--workbook", type=Path)
    parser.add_argument("--json-output", type=Path)
    args = parser.parse_args()
    project_root = args.project_root.resolve()
    workbook_path = (args.workbook or project_root / "assets/registry/ShellStorm2_美术资产台账_v001.xlsx").resolve()
    result = classify(workbook_path, project_root)
    encoded = json.dumps(result, ensure_ascii=False, indent=2)
    if args.json_output:
        output_path = args.json_output if args.json_output.is_absolute() else project_root / args.json_output
        output_path.parent.mkdir(parents=True, exist_ok=True)
        output_path.write_text(encoded + "\n", encoding="utf-8")
    print(json.dumps({key: value for key, value in result.items() if key not in {"records", "unresolved"}}, ensure_ascii=False, indent=2))
    if result["unresolved_parent_count"]:
        print("ASSET_REPOSITORY_CLASSIFICATION_FAILED", file=sys.stderr)
        return 1
    print(f"ASSET_REPOSITORY_CLASSIFICATION_OK candidates={result['classified_candidate_count']}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
