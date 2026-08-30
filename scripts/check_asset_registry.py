#!/usr/bin/env python3
"""Validate the authoritative ShellStorm2 art-asset workbook.

Structure mode is the AST-01 gate (IDs, enums, formulas and dedupe keys). Full mode
also reports path, SHA and production-filename debt for AST-02/03. The checker never
updates the ledger: every hash drift still requires classification and asset QA.
"""

from __future__ import annotations

import argparse
import hashlib
import json
import re
import sys
from collections import Counter, defaultdict
from pathlib import Path
from typing import Any

try:
    from openpyxl import load_workbook
except ImportError as exc:  # pragma: no cover - actionable setup failure
    raise SystemExit("openpyxl is required to check the asset registry: pip install openpyxl") from exc


ALLOWED_CATEGORIES = {"角色", "敌人", "武器", "道具", "场景", "场景道具", "UI", "特效", "音频"}
ALLOWED_STATUSES = {"待制作", "程序占位", "已完成", "原型已接入", "弃用"}
ALLOWED_PRIORITIES = {"P0", "P1", "P2"}
ACTIVE_STATUSES = {"已完成", "原型已接入"}
ASSET_ID_PATTERN = re.compile(r"^[A-Z0-9]+(?:-[A-Z0-9]+)+$")
PRODUCTION_NAME_PATTERN = re.compile(
    r"^[a-z0-9]+(?:_[a-z0-9]+)*_v[0-9]{3}\.(?:png|webp|svg|wav|ogg|glb|gltf|tres|tscn|blend)$"
)
PATH_SPLIT_PATTERN = re.compile(r"[;；\n]+")
CANONICAL_FILENAME_EXCEPTIONS = {
    "assets/art/shared/palette/设施低亮多巴胺色盘_10x10_512.png",
}


def _text(value: Any) -> str:
    return "" if value is None else str(value).strip()


def _dedupe_key(row: tuple[Any, ...]) -> str:
    return "|".join(_text(row[index]).lower() for index in (2, 3, 4, 5, 7, 8))


def _sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _registry_paths(raw: str, project_root: Path) -> list[Path]:
    paths: list[Path] = []
    for token in PATH_SPLIT_PATTERN.split(raw):
        token = token.strip()
        if not token or token.startswith("res://"):
            token = token.removeprefix("res://")
        if not token or any(marker in token for marker in ("待", "无", "N/A")):
            continue
        candidate = Path(token)
        if not candidate.is_absolute():
            candidate = project_root / candidate
        paths.append(candidate)
    return paths


def check_registry(workbook_path: Path, project_root: Path, scope: str) -> dict[str, Any]:
    workbook = load_workbook(workbook_path, data_only=False, read_only=False)
    if "资产主表" not in workbook.sheetnames or "总览" not in workbook.sheetnames:
        return {"fatal": ["missing required sheets: 资产主表 / 总览"]}
    main = workbook["资产主表"]
    overview = workbook["总览"]
    rows = [tuple(main.iter_rows(min_row=row, max_row=row, min_col=1, max_col=24, values_only=True))[0]
            for row in range(6, main.max_row + 1) if _text(main.cell(row=row, column=1).value)]

    issues: dict[str, list[dict[str, Any]]] = defaultdict(list)
    asset_ids: Counter[str] = Counter()
    dedupe_rows: dict[str, list[int]] = defaultdict(list)
    for offset, row in enumerate(rows, start=6):
        asset_id = _text(row[0])
        category = _text(row[2])
        status = _text(row[10])
        priority = _text(row[11])
        asset_ids[asset_id] += 1
        dedupe_rows[_dedupe_key(row)].append(offset)
        if not ASSET_ID_PATTERN.fullmatch(asset_id):
            issues["invalid_asset_id"].append({"row": offset, "asset_id": asset_id})
        if category not in ALLOWED_CATEGORIES:
            issues["invalid_category"].append({"row": offset, "asset_id": asset_id, "value": category})
        if status not in ALLOWED_STATUSES:
            issues["invalid_status"].append({"row": offset, "asset_id": asset_id, "value": status})
        if priority not in ALLOWED_PRIORITIES:
            issues["invalid_priority"].append({"row": offset, "asset_id": asset_id, "value": priority})
        if not _text(main.cell(offset, 18).value).startswith("="):
            issues["missing_dedupe_formula"].append({"row": offset, "column": "R", "asset_id": asset_id})
        if not _text(main.cell(offset, 19).value).startswith("="):
            issues["missing_dedupe_formula"].append({"row": offset, "column": "S", "asset_id": asset_id})

        if scope != "full":
            continue
        raw_path = _text(row[14])
        raw_sha = _text(row[19]).lower()
        paths = _registry_paths(raw_path, project_root)
        if status != "弃用":
            for path in paths:
                try:
                    relative_path = path.relative_to(project_root)
                except ValueError:
                    relative_path = path
                if (
                    path.is_file()
                    and str(relative_path).startswith("assets/art/")
                    and relative_path.as_posix() not in CANONICAL_FILENAME_EXCEPTIONS
                    and not PRODUCTION_NAME_PATTERN.fullmatch(path.name)
                ):
                    issues["noncanonical_filename"].append(
                        {"row": offset, "asset_id": asset_id, "path": str(relative_path)}
                    )
        if status not in ACTIVE_STATUSES:
            continue
        if not raw_path:
            issues["missing_path"].append({"row": offset, "asset_id": asset_id})
        if not raw_sha:
            issues["missing_sha"].append({"row": offset, "asset_id": asset_id})
        existing_paths = [path for path in paths if path.is_file()]
        for path in paths:
            if not path.exists():
                issues["path_not_found"].append({"row": offset, "asset_id": asset_id, "path": str(path)})
        if len(existing_paths) == 1 and raw_sha and re.fullmatch(r"[0-9a-f]{64}", raw_sha):
            actual_sha = _sha256(existing_paths[0])
            if actual_sha != raw_sha:
                issues["sha_mismatch"].append(
                    {"row": offset, "asset_id": asset_id, "path": str(existing_paths[0]), "recorded": raw_sha, "actual": actual_sha}
                )

    for asset_id, count in asset_ids.items():
        if count > 1:
            issues["duplicate_asset_id"].append({"asset_id": asset_id, "count": count})
    for key, row_numbers in dedupe_rows.items():
        if key and len(row_numbers) > 1:
            issues["duplicate_dedupe_key"].append({"key": key, "rows": row_numbers})

    expected_end = 5 + len(rows)
    required_overview_formulas = ("A6", "C6", "E6", "G6")
    for cell_ref in required_overview_formulas:
        formula = _text(overview[cell_ref].value)
        if not formula.startswith("=") or f"${expected_end}" not in formula:
            issues["stale_overview_formula"].append({"cell": cell_ref, "formula": formula, "expected_end": expected_end})
    for row_number in range(10, 19):
        for column in ("B", "C"):
            cell_ref = f"{column}{row_number}"
            formula = _text(overview[cell_ref].value)
            if not formula.startswith("=") or f"${expected_end}" not in formula:
                issues["stale_overview_formula"].append({"cell": cell_ref, "formula": formula, "expected_end": expected_end})

    return {
        "workbook": str(workbook_path),
        "scope": scope,
        "asset_count": len(rows),
        "status_counts": Counter(_text(row[10]) for row in rows),
        "category_counts": Counter(_text(row[2]) for row in rows),
        "issue_counts": {key: len(value) for key, value in sorted(issues.items())},
        "issues": dict(sorted(issues.items())),
    }


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--project-root", type=Path, default=Path(__file__).resolve().parents[1])
    parser.add_argument("--workbook", type=Path)
    parser.add_argument("--scope", choices=("structure", "full"), default="full")
    parser.add_argument("--json-output", type=Path)
    args = parser.parse_args()
    project_root = args.project_root.resolve()
    workbook_path = (args.workbook or project_root / "assets/registry/ShellStorm2_美术资产台账_v001.xlsx").resolve()
    result = check_registry(workbook_path, project_root, args.scope)
    encoded = json.dumps(result, ensure_ascii=False, indent=2, default=dict)
    if args.json_output:
        args.json_output.parent.mkdir(parents=True, exist_ok=True)
        args.json_output.write_text(encoded + "\n", encoding="utf-8")
    print(encoded)
    issue_count = sum(result.get("issue_counts", {}).values()) + len(result.get("fatal", []))
    if issue_count:
        print(f"ASSET_REGISTRY_CHECK_FAILED scope={args.scope} issues={issue_count}", file=sys.stderr)
        return 1
    print(f"ASSET_REGISTRY_CHECK_OK scope={args.scope} assets={result.get('asset_count', 0)}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
