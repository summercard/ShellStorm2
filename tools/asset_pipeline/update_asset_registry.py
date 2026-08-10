"""Register the independent Blender facility and weapon assets in the project workbook."""

from __future__ import annotations

import copy
import hashlib
import json
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook


PROJECT = Path(__file__).resolve().parents[2]
REGISTRY = next((PROJECT / "assets/registry").glob("*.xlsx"))
MANIFEST = json.loads((PROJECT / "assets/art/asset_import_manifest_v001.json").read_text(encoding="utf-8"))
TODAY = datetime(2026, 8, 10)

ACTIVE_WEAPONS = {
    "hair_dryer": ("WPN-GUN-BP-PISTOL", "bp_pistol", "豌豆手枪 / 吹风机造型"),
    "double_barrel_cannon": ("WPN-GUN-BP-SHOTGUN", "bp_shotgun", "散射喷壶 / 双管炮造型"),
    "broom_rifle": ("WPN-GUN-BP-RIFLE", "bp_rifle", "步枪 / 扫帚步枪造型"),
    "water_tank_blaster": ("WPN-GUN-BP-MACHINEGUN", "bp_machinegun", "蜂窝机枪 / 水箱爆能枪造型"),
    "candy_sniper": ("WPN-GUN-BP-SNIPER", "bp_sniper", "弹弓狙击 / 糖果狙击枪造型"),
    "toaster_launcher": ("WPN-GUN-BP-LAUNCHER", "bp_launcher", "反胃榴弹筒 / 烤面包机造型"),
    "gumball_cannon": ("WPN-GUN-BP-CHARGE", "bp_charge", "蓄力萝卜炮 / 口香糖机造型"),
}

FUTURE_WEAPON_IDS = {
    "megaphone_cannon": "WPN-GUN-FUTURE-MEGAPHONE-CANNON-3D",
    "guitar_blaster": "WPN-GUN-FUTURE-GUITAR-BLASTER-3D",
    "spatula_rifle": "WPN-GUN-FUTURE-SPATULA-RIFLE-3D",
    "frying_pan_cannon": "WPN-GUN-FUTURE-FRYING-PAN-CANNON-3D",
    "scope_cannon": "WPN-GUN-FUTURE-SCOPE-CANNON-3D",
    "popcorn_blaster": "WPN-GUN-FUTURE-POPCORN-BLASTER-3D",
    "soda_straw_blaster": "WPN-GUN-FUTURE-SODA-STRAW-BLASTER-3D",
    "crocodile_cannon": "WPN-GUN-FUTURE-CROCODILE-CANNON-3D",
    "camera_blaster": "WPN-GUN-FUTURE-CAMERA-BLASTER-3D",
    "tissue_box_cannon": "WPN-GUN-FUTURE-TISSUE-BOX-CANNON-3D",
    "fan_blaster": "WPN-GUN-FUTURE-FAN-BLASTER-3D",
}


def sha256(resource_path: str) -> str:
    path = PROJECT / resource_path
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def runtime_path(kind: str, slug: str) -> str:
    if kind == "weapon":
        return f"assets/art/weapons/weapon_3d/runtime/{slug}/wpn_{slug}_root_top3d_v001.tscn"
    version = "v002" if slug == "vending_machine" else "v001"
    return f"assets/art/props/base_world_3d/runtime/{slug}/prp_base_{slug}_root_top3d_{version}.tscn"


def update_row(ws, row: int, values: dict[int, object]) -> None:
    for column, value in values.items():
        ws.cell(row, column).value = value
    ws.cell(row, 18).value = f'=LOWER(TRIM(C{row})&"|"&TRIM(D{row})&"|"&TRIM(E{row})&"|"&TRIM(F{row})&"|"&TRIM(H{row})&"|"&TRIM(I{row}))'


def append_row(ws, template_row: int, values: list[object]) -> int:
    asset_id = str(values[0])
    row = next(
        (index for index in range(6, ws.max_row + 1) if str(ws.cell(index, 1).value or "") == asset_id),
        0,
    )
    if row == 0:
        row = max(index for index in range(6, ws.max_row + 1) if ws.cell(index, 1).value) + 1
    for column in range(1, 25):
        source = ws.cell(template_row, column)
        target = ws.cell(row, column)
        if source.has_style:
            target._style = copy.copy(source._style)
        if source.number_format:
            target.number_format = source.number_format
    for column, value in enumerate(values, start=1):
        ws.cell(row, column).value = value
    ws.cell(row, 18).value = f'=LOWER(TRIM(C{row})&"|"&TRIM(D{row})&"|"&TRIM(E{row})&"|"&TRIM(F{row})&"|"&TRIM(H{row})&"|"&TRIM(I{row}))'
    return row


def main() -> None:
    workbook = load_workbook(REGISTRY)
    ws = workbook.worksheets[1]
    rows = {str(ws.cell(row, 1).value): row for row in range(6, ws.max_row + 1) if ws.cell(row, 1).value}

    for slug, (asset_id, logic_id, display_name) in ACTIVE_WEAPONS.items():
        data = MANIFEST["weapons"][slug]
        scene_path = runtime_path("weapon", slug)
        update_row(ws, rows[asset_id], {
            2: display_name,
            8: "俯视3D / local -Z 枪口",
            9: "default / 标准握把与8类挂点",
            10: "手持/背负/地面/UI 共用独立PackedScene",
            11: "正式美术已接入",
            14: "握把原点；local -Z枪口；3材质；独立GLB/PackedScene",
            15: scene_path,
            16: f"{data['source']}; src/combat3d/WeaponModel3D.gd",
            17: f"{logic_id}; {data['name_cn']}; 握把原点; -Z枪口; 通用挂点; 多巴胺配色",
            19: "唯一",
            20: sha256(data["glb"]),
            21: "Codex",
            22: TODAY,
            23: "用户自制Blender资产",
            24: "由18枪合集拆成独立资产；根节点即握把原点，枪口统一local -Z；无需玩家侧逐枪旋转补丁。",
        })

    for slug, asset_id in FUTURE_WEAPON_IDS.items():
        data = MANIFEST["weapons"][slug]
        scene_path = runtime_path("weapon", slug)
        append_row(ws, 24, [
            asset_id, data["name_cn"], "武器", "gunbody", f"future_{slug}", "root", None,
            "俯视3D / local -Z 枪口", "default / 标准握把与8类挂点",
            "未来枪械池/独立资产", "正式美术已入库/待玩法映射", "P2", "v001",
            "握把原点；local -Z枪口；3材质；独立GLB/PackedScene", scene_path,
            data["source"], f"{slug}; {data['name_cn']}; 未来枪械; 通用挂点; 多巴胺配色", None,
            "唯一", sha256(data["glb"]), "Codex", TODAY, "用户自制Blender资产",
            "暂不绑定现有蓝图ID；已独立导入并统一Grip/Muzzle/Scope/Magazine/Stock/Tactical等节点。",
        ])

    facility_updates = {
        "mission_operations": ("PRP-BASE-BRIEFING", "远征情报终端 / 战术情报指挥桌", "mission_operations"),
        "weapon_workshop": ("PRP-BASE-WORKSHOP", "枪械工坊 / 赛博维修工作台", "weapon_workshop"),
        "vending_machine": ("PRP-BASE-VENDING-MACHINE-3D", "基地科幻自动贩卖机", "base_vending"),
    }
    for slug, (asset_id, display_name, logic_id) in facility_updates.items():
        data = MANIFEST["facilities"][slug]
        update_row(ws, rows[asset_id], {
            2: display_name,
            3: "3D道具",
            4: "base_facility",
            5: logic_id,
            8: "Top3D / local -Z 正面",
            9: "default / 独立交互场景",
            10: "BaseWorld3D/TowerDescent99F",
            11: "正式美术已接入",
            13: "v003" if slug in ("mission_operations", "weapon_workshop") else "v002",
            14: "4材质；主体/自发光分离；独立GLB/PackedScene",
            15: runtime_path("facility", slug),
            16: f"{data['source']}; src/base3d/BaseFacility3D.gd; src/world3d/TowerDescent3D.gd",
            17: f"{logic_id}; {data['name_cn']}; 基地设施; 多巴胺配色; 自发光分离",
            19: "唯一",
            20: sha256(data["glb"]),
            21: "Codex",
            22: TODAY,
            23: "用户自制Blender资产",
            24: "正式模型替换程序占位；保留稳定facility_id和交互协议。",
        })

    for slug, asset_id, display_name in (
        ("workshop_stool", "PRP-BASE-WORKSHOP-STOOL-3D", "赛博维修圆凳"),
        ("mission_command_chair", "PRP-BASE-MISSION-COMMAND-CHAIR-3D", "战术指挥椅"),
    ):
        data = MANIFEST["decor_props"][slug]
        append_row(ws, rows["PRP-BASE-VENDING-MACHINE-3D"], [
            asset_id, display_name, "3D道具", "decor_prop", slug, "root", None,
            "Top3D / local -Z 正面", "default / 独立装饰场景", "基地99层美术布置层",
            "正式美术已接入", "P2", "v001", "最多4材质；独立GLB/PackedScene；无交互碰撞",
            data["runtime"], data["source"], f"{slug}; {display_name}; 独立座椅; 可移动装饰",
            None, "唯一", sha256(data["glb"]), "Codex", TODAY, "用户自制Blender资产",
            "已从对应设施主体拆出；在场景装饰集合中独立摆放，不绑定facility_id。",
        ])

    for slug, asset_id, logic_id in (
        ("locker_station", "PRP-BASE-LOCKER-STATION-3D", "future_locker_station"),
        ("retro_tv_station", "PRP-BASE-RETRO-TV-STATION-3D", "future_retro_tv_station"),
    ):
        data = MANIFEST["facilities"][slug]
        append_row(ws, rows["PRP-BASE-VENDING-MACHINE-3D"], [
            asset_id, data["name_cn"], "3D道具", "base_facility", logic_id, "root", "PRP-BASE-FACILITY-3D",
            "Top3D / local -Z 正面", "default / 独立交互场景", "未来基地设施池", "正式美术已入库/待玩法映射",
            "P2", "v002", "4材质；主体/自发光分离；独立GLB/PackedScene", runtime_path("facility", slug),
            data["source"], f"{logic_id}; {data['name_cn']}; 基地设施; 未开放", None, "唯一",
            sha256(data["glb"]), "Codex", TODAY, "用户自制Blender资产",
            "v002按5米地板模数整体等比例强化；仅作为墙边装饰入库，不绑定当前未开放设施功能。",
        ])

    version_ws = workbook.worksheets[7]
    version_ws.cell(12, 1).value = "v0.1.9"
    version_ws.cell(12, 2).value = TODAY
    version_ws.cell(12, 3).value = "正式基地设施与独立枪械美术接入"
    version_ws.cell(12, 4).value = "5设施/18枪/统一挂点"
    version_ws.cell(12, 5).value = "基地接入贩卖机、情报台、工作台；18把枪拆为独立资产，7把映射现有蓝图，其余入未来资产池。"
    version_ws.cell(12, 6).value = "不改玩法稳定ID；枪械统一Grip原点/local -Z枪口与8类挂点；旧程序模型保留为回退。"
    version_ws.cell(12, 7).value = "Codex"

    workbook.save(REGISTRY)
    print(f"Updated registry: {REGISTRY}")


if __name__ == "__main__":
    main()
